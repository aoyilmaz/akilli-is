"""
Akıllı İş - PDKS (Personel Devam Kontrol Sistemi) Servisi

ZKTeco API, CSV ve Excel formatlarından giriş-çıkış verilerini import eder.
"""

import csv
import io
from datetime import datetime, date, time
from decimal import Decimal
from typing import List, Dict, Optional, Tuple, BinaryIO, Union
from dataclasses import dataclass
from enum import Enum
import json
import requests
from pathlib import Path

try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from sqlalchemy import or_
from sqlalchemy.orm import Session
from database.base import get_session
from database.models.hr import Employee, Attendance, AttendanceStatus


class PDKSSource(Enum):
    """PDKS veri kaynağı türleri"""

    ZKTECO_API = "zkteco_api"
    CSV = "csv"
    EXCEL = "excel"


@dataclass
class AttendanceRecord:
    """Puantaj kaydı veri sınıfı"""

    employee_no: str
    date: date
    check_in: Optional[datetime] = None
    check_out: Optional[datetime] = None
    status: AttendanceStatus = AttendanceStatus.PRESENT
    source: PDKSSource = PDKSSource.CSV
    raw_data: Optional[Dict] = None


class ZKTecoClient:
    """
    ZKTeco PDKS cihazı API istemcisi

    ZKTeco cihazlarının REST API'si ile iletişim kurar.
    Desteklenen cihazlar: ZKTeco K40, K50, iClock serisi vb.
    """

    def __init__(self, host: str, port: int = 80, api_key: str = None):
        """
        Args:
            host: Cihaz IP adresi veya hostname
            port: API portu (varsayılan 80)
            api_key: Opsiyonel API anahtarı
        """
        self.base_url = f"http://{host}:{port}"
        self.api_key = api_key
        self.session = requests.Session()
        if api_key:
            self.session.headers["Authorization"] = f"Bearer {api_key}"

    def test_connection(self) -> bool:
        """Bağlantı testi"""
        try:
            response = self.session.get(f"{self.base_url}/api/status", timeout=5)
            return response.status_code == 200
        except requests.RequestException:
            return False

    def get_attendance_logs(
        self, start_date: date, end_date: date, employee_ids: List[str] = None
    ) -> List[Dict]:
        """
        Belirtilen tarih aralığındaki giriş-çıkış loglarını çeker

        Args:
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            employee_ids: Opsiyonel çalışan ID listesi

        Returns:
            Giriş-çıkış kayıtları listesi
        """
        params = {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
        }
        if employee_ids:
            params["employee_ids"] = ",".join(employee_ids)

        try:
            response = self.session.get(
                f"{self.base_url}/api/attendance/logs", params=params, timeout=30
            )
            response.raise_for_status()
            return response.json().get("data", [])
        except requests.RequestException as e:
            raise PDKSConnectionError(f"ZKTeco API bağlantı hatası: {e}")

    def get_employees(self) -> List[Dict]:
        """Cihazda kayıtlı çalışanları çeker"""
        try:
            response = self.session.get(f"{self.base_url}/api/employees", timeout=10)
            response.raise_for_status()
            return response.json().get("data", [])
        except requests.RequestException as e:
            raise PDKSConnectionError(f"ZKTeco API bağlantı hatası: {e}")


class PDKSConnectionError(Exception):
    """PDKS bağlantı hatası"""

    pass


class PDKSImportError(Exception):
    """PDKS veri import hatası"""

    pass


class PDKSService:
    """
    PDKS (Personel Devam Kontrol Sistemi) Servisi

    ZKTeco API, CSV ve Excel formatlarından giriş-çıkış verilerini
    import edip Attendance kayıtlarına dönüştürür.
    """

    # CSV kolon eşleştirmeleri (esnek import için)
    CSV_COLUMN_MAPPINGS = {
        "employee_no": ["sicil_no", "employee_no", "sicil", "personel_no", "id"],
        "date": ["tarih", "date", "gun"],
        "check_in": ["giris", "check_in", "giris_saati", "in_time"],
        "check_out": ["cikis", "check_out", "cikis_saati", "out_time"],
    }

    def __init__(self, session: Session = None):
        self.session = session or get_session()
        self._employee_cache: Dict[str, Employee] = {}

    def _get_employee(self, employee_no: str) -> Optional[Employee]:
        """Çalışanı cache'den veya DB'den getir"""
        if employee_no not in self._employee_cache:
            emp = (
                self.session.query(Employee)
                .filter(Employee.employee_no == employee_no, Employee.is_active == True)
                .first()
            )
            self._employee_cache[employee_no] = emp
        return self._employee_cache.get(employee_no)

    def _clear_cache(self):
        """Cache'i temizle"""
        self._employee_cache.clear()

    # ========== ZKTeco API Import ==========

    def import_from_zkteco(
        self,
        host: str,
        port: int = 80,
        api_key: str = None,
        start_date: date = None,
        end_date: date = None,
    ) -> Tuple[int, int, List[str]]:
        """
        ZKTeco cihazından verileri import et

        Args:
            host: Cihaz IP/hostname
            port: API portu
            api_key: API anahtarı
            start_date: Başlangıç tarihi (varsayılan: bu ayın başı)
            end_date: Bitiş tarihi (varsayılan: bugün)

        Returns:
            (başarılı_sayı, hatalı_sayı, hata_mesajları)
        """
        if start_date is None:
            today = date.today()
            start_date = date(today.year, today.month, 1)
        if end_date is None:
            end_date = date.today()

        client = ZKTecoClient(host, port, api_key)

        # Bağlantı testi
        if not client.test_connection():
            raise PDKSConnectionError(f"ZKTeco cihazına bağlanılamadı: {host}:{port}")

        # Logları çek
        logs = client.get_attendance_logs(start_date, end_date)

        # Kayıtlara dönüştür
        records = []
        for log in logs:
            try:
                record = AttendanceRecord(
                    employee_no=log.get("employee_id", ""),
                    date=datetime.fromisoformat(log["date"]).date(),
                    check_in=(
                        datetime.fromisoformat(log["check_in"])
                        if log.get("check_in")
                        else None
                    ),
                    check_out=(
                        datetime.fromisoformat(log["check_out"])
                        if log.get("check_out")
                        else None
                    ),
                    source=PDKSSource.ZKTECO_API,
                    raw_data=log,
                )
                records.append(record)
            except (KeyError, ValueError) as e:
                continue

        return self._save_records(records)

    # ========== CSV Import ==========

    def import_from_csv(
        self,
        file_content: Union[str, BinaryIO],
        delimiter: str = ",",
        encoding: str = "utf-8",
    ) -> Tuple[int, int, List[str]]:
        """
        CSV dosyasından verileri import et

        Args:
            file_content: CSV içeriği (string veya file object)
            delimiter: Ayraç karakteri
            encoding: Karakter kodlaması

        Returns:
            (başarılı_sayı, hatalı_sayı, hata_mesajları)
        """
        # String veya file object kontrolü
        if hasattr(file_content, "read"):
            content = file_content.read()
            if isinstance(content, bytes):
                content = content.decode(encoding)
        else:
            content = file_content

        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)

        # Kolon eşleştirmelerini bul
        fieldnames = (
            [f.lower().strip() for f in reader.fieldnames] if reader.fieldnames else []
        )
        column_map = self._find_column_mappings(fieldnames)

        if not column_map.get("employee_no") or not column_map.get("date"):
            raise PDKSImportError(
                "CSV'de gerekli kolonlar bulunamadı. "
                "Beklenen: sicil_no/employee_no, tarih/date"
            )

        records = []
        errors = []

        for row_num, row in enumerate(reader, start=2):
            try:
                record = self._parse_csv_row(row, column_map, row_num)
                if record:
                    records.append(record)
            except Exception as e:
                errors.append(f"Satır {row_num}: {str(e)}")

        success, failed, save_errors = self._save_records(records)
        return success, failed + len(errors), errors + save_errors

    def _find_column_mappings(self, fieldnames: List[str]) -> Dict[str, str]:
        """CSV kolonlarını eşleştir"""
        result = {}
        for key, aliases in self.CSV_COLUMN_MAPPINGS.items():
            for alias in aliases:
                if alias in fieldnames:
                    result[key] = alias
                    break
        return result

    def _parse_csv_row(
        self, row: Dict, column_map: Dict[str, str], row_num: int
    ) -> Optional[AttendanceRecord]:
        """CSV satırını parse et"""
        # Sicil no
        emp_col = column_map["employee_no"]
        employee_no = row.get(emp_col, "").strip()
        if not employee_no:
            return None

        # Tarih
        date_col = column_map["date"]
        date_str = row.get(date_col, "").strip()
        record_date = self._parse_date(date_str)
        if not record_date:
            raise ValueError(f"Geçersiz tarih formatı: {date_str}")

        # Giriş saati
        check_in = None
        if column_map.get("check_in"):
            in_str = row.get(column_map["check_in"], "").strip()
            if in_str:
                check_in = self._parse_datetime(record_date, in_str)

        # Çıkış saati
        check_out = None
        if column_map.get("check_out"):
            out_str = row.get(column_map["check_out"], "").strip()
            if out_str:
                check_out = self._parse_datetime(record_date, out_str)

        return AttendanceRecord(
            employee_no=employee_no,
            date=record_date,
            check_in=check_in,
            check_out=check_out,
            source=PDKSSource.CSV,
            raw_data=dict(row),
        )

    def _parse_date(self, date_str: str) -> Optional[date]:
        """Tarih parse et (çoklu format desteği)"""
        formats = [
            "%d.%m.%Y",  # 19.01.2026
            "%d/%m/%Y",  # 19/01/2026
            "%Y-%m-%d",  # 2026-01-19
            "%d-%m-%Y",  # 19-01-2026
        ]
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_datetime(self, record_date: date, time_str: str) -> Optional[datetime]:
        """Saat parse et ve datetime oluştur"""
        formats = [
            "%H:%M:%S",  # 08:30:00
            "%H:%M",  # 08:30
            "%H.%M",  # 08.30
        ]
        for fmt in formats:
            try:
                t = datetime.strptime(time_str, fmt).time()
                return datetime.combine(record_date, t)
            except ValueError:
                continue
        return None

    # ========== Excel Import ==========

    def import_from_excel(
        self, file_path: Union[str, Path, BinaryIO], sheet_name: Optional[str] = None
    ) -> Tuple[int, int, List[str]]:
        """
        Excel dosyasından verileri import et

        Args:
            file_path: Excel dosya yolu veya file object
            sheet_name: Sayfa adı (varsayılan: ilk sayfa)

        Returns:
            (başarılı_sayı, hatalı_sayı, hata_mesajları)
        """
        if not EXCEL_AVAILABLE:
            raise PDKSImportError(
                "Excel desteği için 'openpyxl' paketi gerekli. "
                "Yüklemek için: pip install openpyxl"
            )

        # Dosyayı aç
        if hasattr(file_path, "read"):
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        else:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

        # Sayfa seç
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Başlık satırını al
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise PDKSImportError("Excel dosyası boş")

        headers = [str(h).lower().strip() if h else "" for h in rows[0]]
        column_map = self._find_column_mappings(headers)

        if not column_map.get("employee_no") or not column_map.get("date"):
            raise PDKSImportError(
                "Excel'de gerekli kolonlar bulunamadı. "
                "Beklenen: sicil_no/employee_no, tarih/date"
            )

        records = []
        errors = []

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                # Row'u dict'e çevir
                row_dict = {
                    headers[i]: row[i] for i in range(len(headers)) if i < len(row)
                }
                record = self._parse_excel_row(row_dict, column_map, row_num)
                if record:
                    records.append(record)
            except Exception as e:
                errors.append(f"Satır {row_num}: {str(e)}")

        wb.close()

        success, failed, save_errors = self._save_records(records)
        return success, failed + len(errors), errors + save_errors

    def _parse_excel_row(
        self, row: Dict, column_map: Dict[str, str], row_num: int
    ) -> Optional[AttendanceRecord]:
        """Excel satırını parse et"""
        # Sicil no
        emp_col = column_map["employee_no"]
        employee_no = row.get(emp_col)
        if employee_no is None:
            return None
        employee_no = str(employee_no).strip()
        if not employee_no:
            return None

        # Tarih (Excel'de datetime olabilir)
        date_col = column_map["date"]
        date_val = row.get(date_col)
        if isinstance(date_val, datetime):
            record_date = date_val.date()
        elif isinstance(date_val, date):
            record_date = date_val
        elif isinstance(date_val, str):
            record_date = self._parse_date(date_val)
        else:
            raise ValueError(f"Geçersiz tarih: {date_val}")

        if not record_date:
            raise ValueError(f"Tarih parse edilemedi: {date_val}")

        # Giriş saati
        check_in = None
        if column_map.get("check_in"):
            in_val = row.get(column_map["check_in"])
            check_in = self._parse_excel_time(record_date, in_val)

        # Çıkış saati
        check_out = None
        if column_map.get("check_out"):
            out_val = row.get(column_map["check_out"])
            check_out = self._parse_excel_time(record_date, out_val)

        return AttendanceRecord(
            employee_no=employee_no,
            date=record_date,
            check_in=check_in,
            check_out=check_out,
            source=PDKSSource.EXCEL,
            raw_data=row,
        )

    def _parse_excel_time(self, record_date: date, time_val) -> Optional[datetime]:
        """Excel saat değerini parse et"""
        if time_val is None:
            return None

        if isinstance(time_val, datetime):
            return datetime.combine(record_date, time_val.time())
        elif isinstance(time_val, time):
            return datetime.combine(record_date, time_val)
        elif isinstance(time_val, str):
            return self._parse_datetime(record_date, time_val)

        return None

    # ========== Kayıt Kaydetme ==========

    def _save_records(
        self, records: List[AttendanceRecord]
    ) -> Tuple[int, int, List[str]]:
        """
        Kayıtları veritabanına kaydet

        Returns:
            (başarılı_sayı, hatalı_sayı, hata_mesajları)
        """
        success = 0
        failed = 0
        errors = []

        for record in records:
            try:
                # Çalışanı kontrol et
                employee = self._get_employee(record.employee_no)
                if not employee:
                    errors.append(f"{record.employee_no}: Çalışan bulunamadı")
                    failed += 1
                    continue

                # Mevcut kaydı kontrol et
                existing = (
                    self.session.query(Attendance)
                    .filter(
                        Attendance.employee_id == employee.id,
                        Attendance.date == record.date,
                    )
                    .first()
                )

                if existing:
                    # Güncelle
                    if record.check_in:
                        existing.check_in = record.check_in
                    if record.check_out:
                        existing.check_out = record.check_out

                    # Çalışma süresini hesapla
                    if existing.check_in and existing.check_out:
                        diff = existing.check_out - existing.check_in
                        existing.work_minutes = int(diff.total_seconds() / 60)

                    existing.status = self._determine_status(existing)
                else:
                    # Yeni kayıt
                    att = Attendance(
                        employee_id=employee.id,
                        date=record.date,
                        check_in=record.check_in,
                        check_out=record.check_out,
                        status=AttendanceStatus.PRESENT,
                    )

                    # Çalışma süresini hesapla
                    if record.check_in and record.check_out:
                        diff = record.check_out - record.check_in
                        att.work_minutes = int(diff.total_seconds() / 60)

                    att.status = self._determine_status(att)
                    self.session.add(att)

                success += 1

            except Exception as e:
                errors.append(f"{record.employee_no} ({record.date}): {str(e)}")
                failed += 1

        # Commit
        try:
            self.session.commit()
        except Exception as e:
            self.session.rollback()
            raise PDKSImportError(f"Veritabanı kayıt hatası: {e}")

        self._clear_cache()
        return success, failed, errors

    def _determine_status(self, attendance: Attendance) -> AttendanceStatus:
        """Devam durumunu belirle (geç kalma, erken çıkış vb.)"""
        # Varsayılan mesai saatleri (ileride ayarlanabilir)
        expected_in = time(8, 0)  # 08:00
        expected_out = time(17, 0)  # 17:00
        late_tolerance = 15  # 15 dakika tolerans

        if not attendance.check_in:
            return AttendanceStatus.ABSENT

        check_in_time = attendance.check_in.time()

        # Geç kalma kontrolü
        expected_in_dt = datetime.combine(attendance.date, expected_in)
        actual_in_dt = datetime.combine(attendance.date, check_in_time)
        late_minutes = (actual_in_dt - expected_in_dt).total_seconds() / 60

        if late_minutes > late_tolerance:
            return AttendanceStatus.LATE

        # Erken çıkış kontrolü
        if attendance.check_out:
            check_out_time = attendance.check_out.time()
            expected_out_dt = datetime.combine(attendance.date, expected_out)
            actual_out_dt = datetime.combine(attendance.date, check_out_time)
            early_minutes = (expected_out_dt - actual_out_dt).total_seconds() / 60

            if early_minutes > late_tolerance:
                return AttendanceStatus.EARLY_LEAVE

        return AttendanceStatus.PRESENT

    # ========== Yardımcı Metodlar ==========

    def calculate_monthly_attendance(self, year: int, month: int):
        """
        Belirtilen ay için puantaj hesapla.
        Hiç kaydı olmayan günler (hafta sonu hariç) devamsız işaretlenir.
        """
        import calendar

        start_date = date(year, month, 1)
        _, last_day = calendar.monthrange(year, month)
        end_date = date(year, month, last_day)

        # Tüm aktif çalışanları al
        employees = (
            self.session.query(Employee)
            .filter(Employee.is_active == True, Employee.exit_date.is_(None))
            .all()
        )

        for emp in employees:
            # Ayın her günü için kontrol et
            for day_num in range(1, last_day + 1):
                curr_date = date(year, month, day_num)
                if curr_date > date.today():
                    break

                # Hafta sonu kontrolü (Pazar = 6)
                if curr_date.weekday() == 6:
                    continue

                # Mevcut kaydı kontrol et
                existing = (
                    self.session.query(Attendance)
                    .filter(
                        Attendance.employee_id == emp.id, Attendance.date == curr_date
                    )
                    .first()
                )

                if not existing:
                    # Kayıt yoksa devamsız olarak ekle
                    att = Attendance(
                        employee_id=emp.id,
                        date=curr_date,
                        status=AttendanceStatus.ABSENT,
                        notes="Sistem tarafından otomatik oluşturuldu (kayıt yok)",
                    )
                    self.session.add(att)

        self.session.commit()
        return True

    def get_monthly_summary(
        self, year: int, month: int, department_id: int = None, search: str = None
    ) -> List[Dict]:
        """
        Aylık puantaj özeti

        Returns:
            Çalışan bazlı özet listesi
        """
        from sqlalchemy import func, and_
        from sqlalchemy.orm import joinedload

        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        # Temel sorgu - Department'ı eager load yap
        query = (
            self.session.query(Employee)
            .options(joinedload(Employee.department))
            .filter(Employee.is_active == True, Employee.exit_date.is_(None))
        )

        if department_id:
            query = query.filter(Employee.department_id == department_id)

        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    Employee.employee_no.ilike(search_term),
                    Employee.first_name.ilike(search_term),
                    Employee.last_name.ilike(search_term),
                )
            )

        employees = query.all()

        if not employees:
            return []

        employee_ids = [e.id for e in employees]

        # Tüm puantajları tek seferde çek
        all_attendances = (
            self.session.query(Attendance)
            .filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date >= start_date,
                Attendance.date < end_date,
            )
            .all()
        )

        # Puantajları çalışan bazında grupla
        attendance_map = {e_id: [] for e_id in employee_ids}
        for att in all_attendances:
            attendance_map[att.employee_id].append(att)

        results = []
        for emp in employees:
            attendances = attendance_map.get(emp.id, [])

            present = sum(
                1 for a in attendances if a.status == AttendanceStatus.PRESENT
            )
            late = sum(1 for a in attendances if a.status == AttendanceStatus.LATE)
            absent = sum(1 for a in attendances if a.status == AttendanceStatus.ABSENT)
            on_leave = sum(
                1 for a in attendances if a.status == AttendanceStatus.ON_LEAVE
            )
            total_minutes = sum(a.work_minutes or 0 for a in attendances)
            overtime = sum(a.overtime_minutes or 0 for a in attendances)

            results.append(
                {
                    "employee_id": emp.id,
                    "employee_no": emp.employee_no,
                    "employee_name": emp.full_name,
                    "department": emp.department.name if emp.department else "-",
                    "present_days": present,
                    "late_days": late,
                    "absent_days": absent,
                    "leave_days": on_leave,
                    "total_work_hours": round(total_minutes / 60, 1),
                    "overtime_hours": round(overtime / 60, 1),
                }
            )

        return results

    def close(self):
        """Session'ı kapat"""
        if self.session:
            self.session.close()
